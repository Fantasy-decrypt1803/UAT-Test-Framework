"""
reports/uat_report.py
Generates a professional UAT sign-off Excel report.
UBS equivalent: formal UAT evidence document submitted before production release.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import os

REPORTS_DIR = Path(__file__).parent.parent / "reports" / "output"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

NAVY      = "1F3864"
BLUE      = "2E75B6"
WHITE     = "FFFFFF"
GREEN     = "E8F5E9"
GREEN_D   = "27AE60"
RED       = "FDECEA"
RED_D     = "C0392B"
AMBER     = "FFF3CD"
GREY      = "F5F5F5"
DGREY     = "D9D9D9"


def _hdr(ws, row, c1, c2, height=22):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = PatternFill("solid", fgColor=NAVY)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height


def _title_row(ws, text, merge, height=32):
    ws.merge_cells(merge)
    cell = ws[merge.split(":")[0]]
    cell.value     = text
    cell.font      = Font(bold=True, size=14, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row = int("".join(filter(str.isdigit, merge.split(":")[0])))
    ws.row_dimensions[row].height = height


def _auto_width(ws, max_w=60):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, max_w)


def _status_fill(status):
    return {
        "PASS": GREEN, "FAIL": RED, "WARN": AMBER, "NOT RUN": GREY
    }.get(status, WHITE)

def _status_font_color(status):
    return {
        "PASS": GREEN_D, "FAIL": RED_D, "WARN": "B7860B", "NOT RUN": "888888"
    }.get(status, "000000")


def generate_uat_report(results: list, scenario: str = "clean") -> str:
    wb    = Workbook()
    today = datetime.today()

    passed   = sum(1 for r in results if r.status == "PASS")
    failed   = sum(1 for r in results if r.status == "FAIL")
    warned   = sum(1 for r in results if r.status == "WARN")
    total    = len(results)
    crit_fail = [r for r in results if r.status == "FAIL" and r.sla_critical]
    decision = "APPROVED" if not crit_fail else "BLOCKED"
    dec_colour = GREEN if decision == "APPROVED" else RED

    # ══════════════════════════════════════════════════════════
    #  SHEET 1: UAT SIGN-OFF SUMMARY
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "UAT Sign-Off"
    ws1.sheet_view.showGridLines = False

    _title_row(ws1, "UAT Sign-Off Report — Financial Data Pipeline", "A1:G1")

    # Sub-header
    ws1.merge_cells("A2:G2")
    ws1["A2"].value     = f"Scenario: {scenario.upper()}   |   Date: {today.strftime('%Y-%m-%d %H:%M')}   |   Release: v{today.strftime('%Y%m%d')}"
    ws1["A2"].font      = Font(italic=True, size=10, color="555555")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 16

    # UAT Decision banner
    ws1.merge_cells("A3:G3")
    ws1["A3"].value     = f"UAT DECISION: {decision}"
    ws1["A3"].font      = Font(bold=True, size=16, color=WHITE)
    ws1["A3"].fill      = PatternFill("solid", fgColor=GREEN_D if decision == "APPROVED" else RED_D)
    ws1["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 38

    # KPI row
    kpis = [
        ("Total Tests", total,   "D6E4F0"),
        ("PASSED",      passed,  GREEN),
        ("FAILED",      failed,  RED   if failed > 0  else GREEN),
        ("WARNINGS",    warned,  AMBER if warned > 0  else GREEN),
        ("Critical Fails", len(crit_fail), RED if crit_fail else GREEN),
        ("Pass Rate",   f"{passed/total*100:.0f}%", GREEN if passed/total >= 0.9 else AMBER),
    ]
    ws1.row_dimensions[5].height = 42
    for i, (label, value, colour) in enumerate(kpis, 1):
        ws1.cell(row=4, column=i, value=label).font = Font(bold=True, size=9, color=NAVY)
        ws1.cell(row=4, column=i).fill = PatternFill("solid", fgColor=colour)
        ws1.cell(row=4, column=i).alignment = Alignment(horizontal="center")
        ws1.cell(row=5, column=i, value=value).font = Font(bold=True, size=22, color=NAVY)
        ws1.cell(row=5, column=i).fill = PatternFill("solid", fgColor=colour)
        ws1.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions[get_column_letter(i)].width = 18

    # Full results table
    hdrs = ["Test ID", "Category", "Test Name", "Severity", "SLA Critical", "Status", "Detail"]
    _hdr(ws1, 7, 1, len(hdrs))
    for j, h in enumerate(hdrs, 1):
        ws1.cell(row=7, column=j, value=h)

    for r_idx, result in enumerate(results, start=8):
        fill = _status_fill(result.status)
        row_data = [
            result.test_id, result.category, result.name,
            result.severity,
            "YES" if result.sla_critical else "NO",
            result.status,
            result.detail,
        ]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if c_idx == 6:  # Status column — colour the text too
                cell.font = Font(bold=True, color=_status_font_color(result.status))
        ws1.row_dimensions[r_idx].height = 26

    _auto_width(ws1)

    # ══════════════════════════════════════════════════════════
    #  SHEET 2: FAILURES & WARNINGS DETAIL
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Failures & Warnings")
    ws2.sheet_view.showGridLines = False
    _title_row(ws2, "Failures & Warnings — Investigation Detail", "A1:F1")

    issues = [r for r in results if r.status in ("FAIL", "WARN")]
    hdrs2  = ["Test ID", "Status", "Severity", "SLA Critical", "Detail", "Affected Items"]
    _hdr(ws2, 2, 1, len(hdrs2))
    for j, h in enumerate(hdrs2, 1):
        ws2.cell(row=2, column=j, value=h)

    if issues:
        for r_idx, result in enumerate(issues, start=3):
            fill = RED if result.status == "FAIL" else AMBER
            row_data = [
                result.test_id, result.status, result.severity,
                "YES" if result.sla_critical else "NO",
                result.detail,
                result.to_dict()["affected"],
            ]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.fill      = PatternFill("solid", fgColor=fill)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws2.row_dimensions[r_idx].height = 35
    else:
        ws2.cell(row=3, column=1, value="✅ No failures or warnings — all tests passed.")
        ws2.cell(row=3, column=1).font = Font(bold=True, color=GREEN_D)

    _auto_width(ws2)

    # ══════════════════════════════════════════════════════════
    #  SHEET 3: TEST CASES BY CATEGORY
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("By Category")
    ws3.sheet_view.showGridLines = False
    _title_row(ws3, "Test Results by Category", "A1:E1")

    categories = ["COMPLETENESS", "ACCURACY", "TIMELINESS", "FORMAT", "INTEGRITY", "REGRESSION"]
    row = 2
    for cat in categories:
        cat_results = [r for r in results if r.category == cat]
        if not cat_results:
            continue

        # Category header
        ws3.merge_cells(f"A{row}:E{row}")
        ws3[f"A{row}"].value     = f"{cat}  ({len(cat_results)} tests)"
        ws3[f"A{row}"].font      = Font(bold=True, size=11, color=WHITE)
        ws3[f"A{row}"].fill      = PatternFill("solid", fgColor=BLUE)
        ws3[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
        ws3.row_dimensions[row].height = 22
        row += 1

        for result in cat_results:
            fill = _status_fill(result.status)
            vals = [result.test_id, result.name, result.severity,
                    result.status, result.detail[:80]]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=row, column=c, value=v)
                cell.fill      = PatternFill("solid", fgColor=fill)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if c == 4:
                    cell.font = Font(bold=True, color=_status_font_color(result.status))
            ws3.row_dimensions[row].height = 24
            row += 1
        row += 1

    _auto_width(ws3)

    # ══════════════════════════════════════════════════════════
    #  SHEET 4: UAT CHECKLIST (SIGN-OFF FORM)
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Sign-Off Checklist")
    ws4.sheet_view.showGridLines = False
    _title_row(ws4, "Production Release Sign-Off Checklist", "A1:C1")

    checklist = [
        ("UAT completed by",          "Gavrish B",                         "Data Analyst"),
        ("Release version",           f"v{today.strftime('%Y%m%d')}",      ""),
        ("Test environment",          "Staging — MySQL asset_dashboard",   ""),
        ("Test execution date",       today.strftime("%Y-%m-%d"),          ""),
        ("Total test cases run",      str(total),                          ""),
        ("Test cases passed",         str(passed),                         ""),
        ("Test cases failed",         str(failed),                         "RED" if failed > 0 else ""),
        ("SLA-critical failures",     str(len(crit_fail)),                 "RED" if crit_fail else ""),
        ("UAT decision",              decision,                            "GREEN" if decision=="APPROVED" else "RED"),
        ("Approved for production",   "YES" if decision=="APPROVED" else "NO — SEE FAILURES", ""),
        ("Sign-off timestamp",        datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""),
        ("Next release date",         "TBD",                               ""),
    ]

    _hdr(ws4, 2, 1, 3)
    ws4.cell(row=2, column=1, value="Item")
    ws4.cell(row=2, column=2, value="Value")
    ws4.cell(row=2, column=3, value="Notes")

    for r_idx, (item, value, note) in enumerate(checklist, start=3):
        colour = GREY if r_idx % 2 == 0 else WHITE
        if note == "RED":   colour = RED
        if note == "GREEN": colour = GREEN
        ws4.cell(row=r_idx, column=1, value=item).font  = Font(bold=True, color=NAVY)
        ws4.cell(row=r_idx, column=1).fill              = PatternFill("solid", fgColor=colour)
        ws4.cell(row=r_idx, column=2, value=value)
        ws4.cell(row=r_idx, column=2).fill              = PatternFill("solid", fgColor=colour)
        ws4.cell(row=r_idx, column=3, value="")
        ws4.cell(row=r_idx, column=3).fill              = PatternFill("solid", fgColor=colour)
        ws4.row_dimensions[r_idx].height                = 22

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 25

    # ── Save ──────────────────────────────────────────────────
    fname = f"UAT_SignOff_{scenario}_{today.strftime('%Y%m%d_%H%M%S')}.xlsx"
    path  = REPORTS_DIR / fname
    wb.save(path)
    print(f"  [REPORT]  Saved → {path}")
    return str(path)


if __name__ == "__main__":
    print("Run pipeline/promote.py to generate UAT report.")
